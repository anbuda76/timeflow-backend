from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from app.db.session import get_db
from app.models.models import (
    Timesheet, TimesheetEntry, TimesheetStatus,
    User, UserRole, Project, VendorCost
)
from app.core.deps import require_manager
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import calendar

router = APIRouter(prefix="/reports", tags=["reports"])

@router.get("/costs")
def monthly_costs(
    year: int = Query(...),
    month: int | None = Query(None),
    project_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager),
):
    is_approved = Timesheet.status == TimesheetStatus.APPROVED
    is_pending = Timesheet.status != TimesheetStatus.APPROVED

    q = (
        db.query(
            TimesheetEntry.project_id,
            User.id.label("user_id"),
            User.first_name,
            User.last_name,
            User.hourly_rate,
            func.sum(
                case((is_approved, TimesheetEntry.hours), else_=0.0)
            ).label("approved_hours"),
            func.sum(
                case((is_pending, TimesheetEntry.hours), else_=0.0)
            ).label("pending_hours"),
        )
        .join(Timesheet, TimesheetEntry.timesheet_id == Timesheet.id)
        .join(User, Timesheet.user_id == User.id)
        .filter(Timesheet.year == year)
    )
    if month:
        q = q.filter(Timesheet.month == month)
    if project_id:
        q = q.filter(TimesheetEntry.project_id == project_id)
    if current_user.role != UserRole.SUPER_ADMIN:
        q = q.filter(User.organization_id == current_user.organization_id)
    if current_user.role == UserRole.MANAGER:
        team_ids = [u.id for u in db.query(User).filter(User.manager_id == current_user.id).all()]
        team_ids.append(current_user.id)
        q = q.filter(User.id.in_(team_ids))

    rows = q.group_by(
        TimesheetEntry.project_id, User.id,
        User.first_name, User.last_name, User.hourly_rate
    ).all()

    by_project = {}
    by_user = {}
    total_approved_hours = 0.0
    total_approved_cost = 0.0
    total_pending_hours = 0.0
    total_pending_cost = 0.0

    for row in rows:
        rate = row.hourly_rate or 0.0
        appr_h = float(row.approved_hours or 0)
        pend_h = float(row.pending_hours or 0)
        appr_cost = appr_h * rate
        pend_cost = pend_h * rate

        total_approved_hours += appr_h
        total_approved_cost += appr_cost
        total_pending_hours += pend_h
        total_pending_cost += pend_cost

        project = db.get(Project, row.project_id)
        p_name = project.name if project else f"Progetto #{row.project_id}"
        c_name = project.client_name if project else None
        b_hours = project.budget_hours if project else None
        b_amount = project.budget_amount if project else None

        if row.project_id not in by_project:
            by_project[row.project_id] = {
                "project_id": row.project_id,
                "project_name": p_name,
                "client_name": c_name,
                "budget_hours": b_hours,
                "budget_amount": b_amount,
                "approved_hours": 0.0,
                "approved_cost": 0.0,
                "pending_hours": 0.0,
                "pending_cost": 0.0,
            }
        by_project[row.project_id]["approved_hours"] += appr_h
        by_project[row.project_id]["approved_cost"] += appr_cost
        by_project[row.project_id]["pending_hours"] += pend_h
        by_project[row.project_id]["pending_cost"] += pend_cost

        if row.user_id not in by_user:
            by_user[row.user_id] = {
                "user_id": row.user_id,
                "user_name": f"{row.first_name} {row.last_name}",
                "hourly_rate": rate,
                "approved_hours": 0.0,
                "approved_cost": 0.0,
                "pending_hours": 0.0,
                "pending_cost": 0.0,
            }
        by_user[row.user_id]["approved_hours"] += appr_h
        by_user[row.user_id]["approved_cost"] += appr_cost
        by_user[row.user_id]["pending_hours"] += pend_h
        by_user[row.user_id]["pending_cost"] += pend_cost

    # ── Costi esterni (vendor costs) per progetto ─────────────────────────────
    vc_q = db.query(VendorCost).filter(
        VendorCost.organization_id == current_user.organization_id,
    )
    if month:
        vc_q = vc_q.filter(
            func.extract("month", VendorCost.cost_date) == month,
            func.extract("year", VendorCost.cost_date) == year,
        )
    else:
        vc_q = vc_q.filter(func.extract("year", VendorCost.cost_date) == year)
    # Costi senza data rientrano sempre
    vc_no_date = db.query(VendorCost).filter(
        VendorCost.organization_id == current_user.organization_id,
        VendorCost.cost_date == None,
    )
    vendor_cost_by_project: dict[int, float] = {}
    for vc in list(vc_q.all()) + list(vc_no_date.all()):
        vendor_cost_by_project[vc.project_id] = vendor_cost_by_project.get(vc.project_id, 0.0) + vc.amount

    projects_list = []
    for p in by_project.values():
        b_h = p["budget_hours"]
        b_a = p["budget_amount"]
        appr_h = round(p["approved_hours"], 2)
        pend_h = round(p["pending_hours"], 2)
        appr_a = round(p["approved_cost"], 2)
        pend_a = round(p["pending_cost"], 2)
        c_h = round(appr_h + pend_h, 2)
        c_a = round(appr_a + pend_a, 2)
        ext_a = round(vendor_cost_by_project.get(p["project_id"], 0.0), 2)
        total_a = round(c_a + ext_a, 2)

        delta_hours = round(b_h - c_h, 2) if b_h else None
        delta_hours_pct = round((b_h - c_h) / b_h * 100, 1) if b_h else None
        delta_amount = round(b_a - total_a, 2) if b_a else None
        delta_amount_pct = round((b_a - total_a) / b_a * 100, 1) if b_a else None

        projects_list.append({
            "project_id": p["project_id"],
            "project_name": p["project_name"],
            "client_name": p["client_name"],
            "budget_hours": b_h,
            "budget_amount": b_a,
            "approved_hours": appr_h,
            "approved_amount": appr_a,
            "pending_hours": pend_h,
            "pending_amount": pend_a,
            "consuntivo_hours": c_h,
            "consuntivo_amount": c_a,
            "vendor_cost": ext_a,
            "total_amount": total_a,
            "delta_hours": delta_hours,
            "delta_hours_pct": delta_hours_pct,
            "delta_amount": delta_amount,
            "delta_amount_pct": delta_amount_pct,
        })

    users_list = []
    for u in by_user.values():
        users_list.append({
            **u,
            "approved_hours": round(u["approved_hours"], 2),
            "approved_cost": round(u["approved_cost"], 2),
            "pending_hours": round(u["pending_hours"], 2),
            "pending_cost": round(u["pending_cost"], 2),
            "hours": round(u["approved_hours"] + u["pending_hours"], 2),
            "cost": round(u["approved_cost"] + u["pending_cost"], 2),
        })

    return {
        "year": year,
        "month": month,
        "total_approved_hours": round(total_approved_hours, 2),
        "total_approved_cost": round(total_approved_cost, 2),
        "total_pending_hours": round(total_pending_hours, 2),
        "total_pending_cost": round(total_pending_cost, 2),
        "total_hours": round(total_approved_hours + total_pending_hours, 2),
        "total_cost": round(total_approved_cost + total_pending_cost, 2),
        "projects": projects_list,
        "users": users_list,
    }


@router.get("/projects/{project_id}/detail")
def project_cost_detail(
    project_id: int,
    year: int | None = Query(None, description="Se omesso, aggrega tutti gli anni"),
    month: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager),
):
    """Dettaglio costi di una commessa: risorse interne (ore x tariffa) + costi fornitori."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Progetto non trovato")
    if current_user.role != UserRole.SUPER_ADMIN and project.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Progetto non trovato")

    is_approved = Timesheet.status == TimesheetStatus.APPROVED
    is_pending = Timesheet.status != TimesheetStatus.APPROVED

    q = (
        db.query(
            User.id.label("user_id"),
            User.first_name,
            User.last_name,
            User.hourly_rate,
            func.sum(case((is_approved, TimesheetEntry.hours), else_=0.0)).label("approved_hours"),
            func.sum(case((is_pending, TimesheetEntry.hours), else_=0.0)).label("pending_hours"),
        )
        .join(Timesheet, TimesheetEntry.timesheet_id == Timesheet.id)
        .join(User, Timesheet.user_id == User.id)
        .filter(TimesheetEntry.project_id == project_id)
    )
    if year:
        q = q.filter(Timesheet.year == year)
    if month:
        q = q.filter(Timesheet.month == month)
    if current_user.role != UserRole.SUPER_ADMIN:
        q = q.filter(User.organization_id == current_user.organization_id)
    if current_user.role == UserRole.MANAGER:
        team_ids = [u.id for u in db.query(User).filter(User.manager_id == current_user.id).all()]
        team_ids.append(current_user.id)
        q = q.filter(User.id.in_(team_ids))

    rows = q.group_by(
        User.id, User.first_name, User.last_name, User.hourly_rate
    ).all()

    users_detail = []
    internal_total = 0.0
    for r in rows:
        rate = r.hourly_rate or 0.0
        appr_h = round(float(r.approved_hours or 0), 2)
        pend_h = round(float(r.pending_hours or 0), 2)
        appr_c = round(appr_h * rate, 2)
        pend_c = round(pend_h * rate, 2)
        total_c = round(appr_c + pend_c, 2)
        internal_total += total_c
        users_detail.append({
            "user_id": r.user_id,
            "user_name": f"{r.first_name} {r.last_name}",
            "hourly_rate": rate,
            "approved_hours": appr_h,
            "pending_hours": pend_h,
            "total_hours": round(appr_h + pend_h, 2),
            "approved_cost": appr_c,
            "pending_cost": pend_c,
            "total_cost": total_c,
        })

    # ── Costi fornitori sulla commessa ────────────────────────────────────────
    from app.models.models import Vendor
    vendor_names = {v.id: v.name for v in db.query(Vendor).filter(
        Vendor.organization_id == project.organization_id
    ).all()}

    vc_q = db.query(VendorCost).filter(VendorCost.project_id == project_id)
    if current_user.role != UserRole.SUPER_ADMIN:
        vc_q = vc_q.filter(VendorCost.organization_id == current_user.organization_id)

    vendor_total = 0.0
    vendor_list = []
    for vc in vc_q.all():
        if vc.cost_date is not None:
            if year and vc.cost_date.year != year:
                continue
            if month and vc.cost_date.month != month:
                continue
        vendor_total += vc.amount
        vendor_list.append({
            "id": vc.id,
            "vendor_id": vc.vendor_id,
            "vendor_name": vendor_names.get(vc.vendor_id, "—"),
            "amount": round(vc.amount, 2),
            "cost_date": vc.cost_date,
            "description": vc.description,
            "category": vc.category,
        })
    vendor_total = round(vendor_total, 2)
    vendor_list.sort(key=lambda x: x["amount"], reverse=True)

    grand_total = round(internal_total + vendor_total, 2)

    # Percentuale di incidenza sul totale commessa
    for u in users_detail:
        u["pct"] = round(u["total_cost"] / grand_total * 100, 1) if grand_total else 0.0
    users_detail.sort(key=lambda x: x["total_cost"], reverse=True)
    for v in vendor_list:
        v["pct"] = round(v["amount"] / grand_total * 100, 1) if grand_total else 0.0

    # ── Timeline: per mese se anno definito, altrimenti per anno ──────────────
    budget = project.budget_amount or 0.0

    tq = (
        db.query(
            Timesheet.year.label("y"),
            Timesheet.month.label("m"),
            func.sum(case((is_approved, TimesheetEntry.hours * User.hourly_rate), else_=0.0)).label("appr_cost"),
            func.sum(case((is_pending, TimesheetEntry.hours * User.hourly_rate), else_=0.0)).label("pend_cost"),
        )
        .join(Timesheet, TimesheetEntry.timesheet_id == Timesheet.id)
        .join(User, Timesheet.user_id == User.id)
        .filter(TimesheetEntry.project_id == project_id)
    )
    # La timeline copre SEMPRE l'intera vita della commessa, a prescindere dai
    # filtri anno/mese: solo così la cumulata è confrontabile col budget totale.
    if current_user.role != UserRole.SUPER_ADMIN:
        tq = tq.filter(User.organization_id == current_user.organization_id)
    trows = tq.group_by(Timesheet.year, Timesheet.month).all()

    # Tutti i costi fornitori della commessa, non filtrati per periodo
    all_vc = db.query(VendorCost).filter(VendorCost.project_id == project_id)
    if current_user.role != UserRole.SUPER_ADMIN:
        all_vc = all_vc.filter(VendorCost.organization_id == current_user.organization_id)
    all_vc = all_vc.all()

    MONTH_LABELS = ['Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic']

    # Periodi con attività: (anno, mese) da timesheet + (anno, mese) da costi datati
    periods = {(r.y, r.m) for r in trows}
    periods |= {(vc.cost_date.year, vc.cost_date.month) for vc in all_vc if vc.cost_date}

    undated_vendor = round(sum(vc.amount for vc in all_vc if vc.cost_date is None), 2)

    timeline = []
    granularity = "none"

    if periods:
        ymin, ymax = min(p[0] for p in periods), max(p[0] for p in periods)
        first = min(periods)
        last = max(periods)
        span_months = (last[0] - first[0]) * 12 + (last[1] - first[1]) + 1
        granularity = "month" if span_months <= 24 else "year"

        buckets: dict = {}
        if granularity == "month":
            keys = []
            y, m = first
            while (y, m) <= last:
                keys.append((y, m))
                m += 1
                if m > 12:
                    m = 1; y += 1
            labels = {k: f"{MONTH_LABELS[k[1] - 1]} {str(k[0])[2:]}" for k in keys}
            for r in trows:
                b = buckets.setdefault((r.y, r.m), {"appr": 0.0, "pend": 0.0, "vend": 0.0})
                b["appr"] += float(r.appr_cost or 0)
                b["pend"] += float(r.pend_cost or 0)
            for vc in all_vc:
                if vc.cost_date is None:
                    continue
                k = (vc.cost_date.year, vc.cost_date.month)
                buckets.setdefault(k, {"appr": 0.0, "pend": 0.0, "vend": 0.0})["vend"] += vc.amount
        else:
            keys = list(range(ymin, ymax + 1))
            labels = {k: str(k) for k in keys}
            for r in trows:
                b = buckets.setdefault(r.y, {"appr": 0.0, "pend": 0.0, "vend": 0.0})
                b["appr"] += float(r.appr_cost or 0)
                b["pend"] += float(r.pend_cost or 0)
            for vc in all_vc:
                if vc.cost_date is None:
                    continue
                buckets.setdefault(vc.cost_date.year, {"appr": 0.0, "pend": 0.0, "vend": 0.0})["vend"] += vc.amount

        cumulative = 0.0
        for k in keys:
            b = buckets.get(k, {"appr": 0.0, "pend": 0.0, "vend": 0.0})
            period_cost = b["appr"] + b["pend"] + b["vend"]
            cumulative += period_cost
            timeline.append({
                "label": labels[k],
                "approved_cost": round(b["appr"], 2),
                "pending_cost": round(b["pend"], 2),
                "vendor_cost": round(b["vend"], 2),
                "period_cost": round(period_cost, 2),
                "cumulative_cost": round(cumulative, 2),
                "budget_line": round(budget, 2) if budget else None,
                "undated": False,
            })

        # Costi fornitori senza data: bucket finale, così la cumulata quadra
        if undated_vendor > 0:
            cumulative += undated_vendor
            timeline.append({
                "label": "n.d.",
                "approved_cost": 0.0,
                "pending_cost": 0.0,
                "vendor_cost": undated_vendor,
                "period_cost": undated_vendor,
                "cumulative_cost": round(cumulative, 2),
                "budget_line": round(budget, 2) if budget else None,
                "undated": True,
            })

    # Costo totale sull'intera vita della commessa (per il confronto col budget)
    lifetime_internal = sum(float(r.appr_cost or 0) + float(r.pend_cost or 0) for r in trows)
    lifetime_vendor = sum(vc.amount for vc in all_vc)
    lifetime_total = round(lifetime_internal + lifetime_vendor, 2)

    delta = round(budget - lifetime_total, 2) if budget else None
    delta_pct = round((budget - lifetime_total) / budget * 100, 1) if budget else None

    return {
        "project_id": project_id,
        "project_name": project.name,
        "client_name": project.client_name,
        "budget_amount": project.budget_amount,
        "is_active": project.is_active,
        "start_date": project.start_date,
        "end_date": project.end_date,
        "note": project.note,
        "year": year,
        "month": month,
        "users": users_detail,
        "vendor_costs": vendor_list,
        "timeline": timeline,
        "timeline_granularity": granularity,
        "undated_vendor_cost": undated_vendor,
        "internal_cost": round(internal_total, 2),
        "vendor_cost": vendor_total,
        "vendor_pct": round(vendor_total / grand_total * 100, 1) if grand_total else 0.0,
        "total_cost": grand_total,
        "lifetime_cost": lifetime_total,
        "lifetime_internal": round(lifetime_internal, 2),
        "lifetime_vendor": round(lifetime_vendor, 2),
        "is_filtered": bool(year or month),
        "delta_amount": delta,
        "delta_amount_pct": delta_pct,
    }


@router.get("/users/{user_id}/detail")
def user_cost_detail(
    user_id: int,
    year: int | None = Query(None, description="Se omesso, aggrega tutti gli anni"),
    month: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager),
):
    """Dettaglio delle commesse su cui una risorsa ha consuntivato nel periodo."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    if current_user.role != UserRole.SUPER_ADMIN and user.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    if current_user.role == UserRole.MANAGER:
        team_ids = [u.id for u in db.query(User).filter(User.manager_id == current_user.id).all()]
        team_ids.append(current_user.id)
        if user_id not in team_ids:
            raise HTTPException(status_code=403, detail="Utente non nel tuo team")

    is_approved = Timesheet.status == TimesheetStatus.APPROVED
    is_pending = Timesheet.status != TimesheetStatus.APPROVED

    q = (
        db.query(
            TimesheetEntry.project_id,
            TimesheetEntry.is_overtime,
            func.sum(case((is_approved, TimesheetEntry.hours), else_=0.0)).label("approved_hours"),
            func.sum(case((is_pending, TimesheetEntry.hours), else_=0.0)).label("pending_hours"),
        )
        .join(Timesheet, TimesheetEntry.timesheet_id == Timesheet.id)
        .filter(Timesheet.user_id == user_id)
    )
    if year:
        q = q.filter(Timesheet.year == year)
    if month:
        q = q.filter(Timesheet.month == month)
    rows = q.group_by(TimesheetEntry.project_id, TimesheetEntry.is_overtime).all()

    rate = user.hourly_rate or 0.0
    projects_detail = []
    system_detail = []
    overtime_hours = 0.0
    overtime_cost = 0.0
    total_cost = 0.0
    total_hours = 0.0

    for r in rows:
        appr_h = round(float(r.approved_hours or 0), 2)
        pend_h = round(float(r.pending_hours or 0), 2)
        tot_h = round(appr_h + pend_h, 2)
        if tot_h == 0:
            continue
        cost = round(tot_h * rate, 2)
        total_hours += tot_h
        total_cost += cost

        # Le entry di straordinario non hanno progetto
        if r.is_overtime or r.project_id is None:
            overtime_hours += tot_h
            overtime_cost += cost
            continue

        project = db.get(Project, r.project_id)
        item = {
            "project_id": r.project_id,
            "project_name": project.name if project else f"Progetto #{r.project_id}",
            "client_name": project.client_name if project else None,
            "is_system": bool(project.is_system) if project else False,
            "approved_hours": appr_h,
            "pending_hours": pend_h,
            "total_hours": tot_h,
            "approved_cost": round(appr_h * rate, 2),
            "pending_cost": round(pend_h * rate, 2),
            "total_cost": cost,
        }
        (system_detail if item["is_system"] else projects_detail).append(item)

    total_cost = round(total_cost, 2)
    total_hours = round(total_hours, 2)

    for item in projects_detail + system_detail:
        item["pct"] = round(item["total_cost"] / total_cost * 100, 1) if total_cost else 0.0
    projects_detail.sort(key=lambda x: x["total_cost"], reverse=True)
    system_detail.sort(key=lambda x: x["total_cost"], reverse=True)

    billable_cost = round(sum(p["total_cost"] for p in projects_detail), 2)
    billable_hours = round(sum(p["total_hours"] for p in projects_detail), 2)
    system_cost = round(sum(p["total_cost"] for p in system_detail), 2)
    system_hours = round(sum(p["total_hours"] for p in system_detail), 2)

    return {
        "user_id": user_id,
        "user_name": f"{user.first_name} {user.last_name}",
        "hourly_rate": rate,
        "year": year,
        "month": month,
        "projects": projects_detail,
        "system_projects": system_detail,
        "overtime_hours": round(overtime_hours, 2),
        "overtime_cost": round(overtime_cost, 2),
        "overtime_pct": round(overtime_cost / total_cost * 100, 1) if total_cost else 0.0,
        "billable_cost": billable_cost,
        "billable_hours": billable_hours,
        "billable_pct": round(billable_cost / total_cost * 100, 1) if total_cost else 0.0,
        "system_cost": system_cost,
        "system_hours": system_hours,
        "total_cost": total_cost,
        "total_hours": total_hours,
    }


@router.get("/monthly-trend")
def monthly_trend(
    year: int = Query(...),
    project_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager),
):
    is_approved = Timesheet.status == TimesheetStatus.APPROVED
    is_pending = Timesheet.status != TimesheetStatus.APPROVED

    q = (
        db.query(
            Timesheet.month,
            TimesheetEntry.project_id,
            func.sum(
                case((is_approved, TimesheetEntry.hours), else_=0.0)
            ).label("approved_hours"),
            func.sum(
                case((is_pending, TimesheetEntry.hours), else_=0.0)
            ).label("pending_hours"),
            func.sum(
                case((is_approved, TimesheetEntry.hours * User.hourly_rate), else_=0.0)
            ).label("approved_cost"),
            func.sum(
                case((is_pending, TimesheetEntry.hours * User.hourly_rate), else_=0.0)
            ).label("pending_cost"),
        )
        .join(Timesheet, TimesheetEntry.timesheet_id == Timesheet.id)
        .join(User, Timesheet.user_id == User.id)
        .filter(
            Timesheet.year == year,
            User.hourly_rate.isnot(None),
        )
    )
    if project_id:
        q = q.filter(TimesheetEntry.project_id == project_id)
    if current_user.role != UserRole.SUPER_ADMIN:
        q = q.filter(User.organization_id == current_user.organization_id)
    if current_user.role == UserRole.MANAGER:
        team_ids = [u.id for u in db.query(User).filter(User.manager_id == current_user.id).all()]
        team_ids.append(current_user.id)
        q = q.filter(User.id.in_(team_ids))

    rows = q.group_by(Timesheet.month, TimesheetEntry.project_id).all()

    projects_data = {}
    for row in rows:
        pid = row.project_id
        if pid not in projects_data:
            project = db.get(Project, pid)
            projects_data[pid] = {
                "project_id": pid,
                "project_name": project.name if project else f"#{pid}",
                "budget_amount": project.budget_amount if project else None,
                "months": {m: {"approved_hours": 0.0, "pending_hours": 0.0, "approved_cost": 0.0, "pending_cost": 0.0} for m in range(1, 13)},
            }
        m = row.month
        projects_data[pid]["months"][m]["approved_hours"] += float(row.approved_hours or 0)
        projects_data[pid]["months"][m]["pending_hours"] += float(row.pending_hours or 0)
        projects_data[pid]["months"][m]["approved_cost"] += float(row.approved_cost or 0)
        projects_data[pid]["months"][m]["pending_cost"] += float(row.pending_cost or 0)

    # ── Costi fornitori per progetto/mese ─────────────────────────────────────
    vq = db.query(VendorCost).filter(func.extract("year", VendorCost.cost_date) == year)
    if current_user.role != UserRole.SUPER_ADMIN:
        vq = vq.filter(VendorCost.organization_id == current_user.organization_id)
    if project_id:
        vq = vq.filter(VendorCost.project_id == project_id)

    for vc in vq.all():
        if vc.cost_date is None:
            continue
        pid = vc.project_id
        if pid not in projects_data:
            project = db.get(Project, pid)
            projects_data[pid] = {
                "project_id": pid,
                "project_name": project.name if project else f"#{pid}",
                "budget_amount": project.budget_amount if project else None,
                "months": {m: {"approved_hours": 0.0, "pending_hours": 0.0, "approved_cost": 0.0, "pending_cost": 0.0} for m in range(1, 13)},
            }
        md = projects_data[pid]["months"][vc.cost_date.month]
        md["vendor_cost"] = md.get("vendor_cost", 0.0) + vc.amount

    result = []
    for pid, pdata in projects_data.items():
        monthly = []
        cumulative_approved = 0.0
        cumulative_pending = 0.0
        budget = pdata["budget_amount"] or 0
        cumulative_vendor = 0.0
        for m in range(1, 13):
            md = pdata["months"][m]
            vend = md.get("vendor_cost", 0.0)
            cumulative_approved += md["approved_cost"]
            cumulative_pending += md["pending_cost"]
            cumulative_vendor += vend
            monthly.append({
                "month": m,
                "approved_hours": round(md["approved_hours"], 2),
                "pending_hours": round(md["pending_hours"], 2),
                "hours": round(md["approved_hours"] + md["pending_hours"], 2),
                "approved_cost": round(md["approved_cost"], 2),
                "pending_cost": round(md["pending_cost"], 2),
                "vendor_cost": round(vend, 2),
                "cost": round(md["approved_cost"] + md["pending_cost"] + vend, 2),
                "cumulative_cost": round(cumulative_approved + cumulative_pending + cumulative_vendor, 2),
                "cumulative_approved": round(cumulative_approved, 2),
                # Il budget è il tetto complessivo della commessa, non una quota
                # mensile: si espone come valore costante più la % consumata.
                "budget_line": round(budget, 2) if budget else None,
                "budget_pct": round((cumulative_approved + cumulative_pending + cumulative_vendor) / budget * 100, 1) if budget else None,
            })
        result.append({
            "project_id": pid,
            "project_name": pdata["project_name"],
            "budget_amount": pdata["budget_amount"],
            "monthly": monthly,
        })

    return result


@router.get("/export-excel")
def export_excel(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager),
):
    MONTH_NAMES = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
                   'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']

    header_fill = PatternFill("solid", start_color="1d4ed8", end_color="1d4ed8")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    subheader_font = Font(bold=True, size=9)
    weekend_fill = PatternFill("solid", start_color="f1f5f9", end_color="f1f5f9")
    total_fill = PatternFill("solid", start_color="eff6ff", end_color="eff6ff")
    center = Alignment(horizontal="center", vertical="center")

    users_q = db.query(User).filter(
        User.organization_id == current_user.organization_id,
        User.is_active == True,
        User.role != UserRole.SUPER_ADMIN,
    )
    if current_user.role == UserRole.MANAGER:
        team_ids = [u.id for u in db.query(User).filter(User.manager_id == current_user.id).all()]
        team_ids.append(current_user.id)
        users_q = users_q.filter(User.id.in_(team_ids))
    users = users_q.all()

    projects = db.query(Project).filter(
        Project.organization_id == current_user.organization_id,
        Project.is_system == False,
        Project.is_active == True,
    ).all()

    system_projects = db.query(Project).filter(
        Project.organization_id == current_user.organization_id,
        Project.is_system == True,
    ).all()
    system_map = {p.name.upper(): p.id for p in system_projects}

    days_in_month = calendar.monthrange(year, month)[1]

    def get_entries_for_user(user_id):
        ts = db.query(Timesheet).filter(
            Timesheet.user_id == user_id,
            Timesheet.year == year,
            Timesheet.month == month,
        ).first()
        if not ts:
            return {}
        entries = db.query(TimesheetEntry).filter(
            TimesheetEntry.timesheet_id == ts.id
        ).all()
        result = {}
        for e in entries:
            day = e.entry_date.day
            if e.project_id not in result:
                result[e.project_id] = {}
            result[e.project_id][day] = e.hours
        return result

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "RIEPILOGO GIORNATE"

    row_offset = 0
    for user in users:
        entries = get_entries_for_user(user.id)
        base_row = row_offset + 1

        ws1.merge_cells(start_row=base_row, start_column=3, end_row=base_row, end_column=3+days_in_month)
        cell = ws1.cell(row=base_row, column=3, value=f"{user.first_name.upper()} {user.last_name.upper()}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

        ws1.cell(row=base_row, column=1, value="ANNO").font = subheader_font
        ws1.cell(row=base_row, column=2, value=year)
        ws1.cell(row=base_row+1, column=1, value="MESE").font = subheader_font
        ws1.cell(row=base_row+1, column=2, value=MONTH_NAMES[month])
        ws1.cell(row=base_row+2, column=3, value="GIORNI").font = subheader_font

        for d in range(1, days_in_month+1):
            col = d + 3
            ws1.cell(row=base_row+2, column=col, value=d).alignment = center
            ws1.cell(row=base_row+2, column=col).font = Font(bold=True, size=9)
            dow = calendar.weekday(year, month, d)
            day_letter = ['L','M','M','G','V','S','D'][dow]
            ws1.cell(row=base_row+3, column=col, value=day_letter).alignment = center
            ws1.cell(row=base_row+3, column=col).font = Font(size=8)
            if dow >= 5:
                for r in range(base_row+2, base_row+9):
                    ws1.cell(row=r, column=col).fill = weekend_fill

        row_labels = [
            ("ORE", None),
            ("STRAORDINARI", system_map.get("STRAORDINARI")),
            ("PERMESSI", system_map.get("PERMESSI")),
            ("MALATTIA", system_map.get("MALATTIA")),
            ("ASSENZE/FERIE", system_map.get("FERIE")),
        ]

        for i, (label, proj_id) in enumerate(row_labels):
            data_row = base_row + 4 + i
            ws1.cell(row=data_row, column=3, value=label).font = Font(bold=True, size=9)

            for d in range(1, days_in_month+1):
                col = d + 3
                if label == "ORE":
                    val = sum(
                        proj_entries.get(d, 0)
                        for p_id, proj_entries in entries.items()
                        if p_id in [p.id for p in projects]
                    )
                elif proj_id:
                    val = entries.get(proj_id, {}).get(d, 0)
                else:
                    val = 0
                if val:
                    ws1.cell(row=data_row, column=col, value=val).alignment = center

            tot_col = days_in_month + 4
            ws1.cell(row=data_row, column=tot_col, value=f'=SUM({get_column_letter(4)}{data_row}:{get_column_letter(tot_col-1)}{data_row})')
            ws1.cell(row=data_row, column=tot_col).fill = total_fill
            ws1.cell(row=data_row, column=tot_col).font = Font(bold=True, size=9)

        row_offset += 10

    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 16
    for d in range(1, 32):
        ws1.column_dimensions[get_column_letter(d+3)].width = 4
    ws1.column_dimensions[get_column_letter(35)].width = 8

    ws2 = wb.create_sheet("RIEPILOGO PROGETTI")
    ws2.cell(row=1, column=1, value=f"{MONTH_NAMES[month]} {year}").font = Font(bold=True, size=11)
    ws2.cell(row=2, column=2, value="n°").font = subheader_font
    ws2.cell(row=2, column=3, value="PROGETTO").font = subheader_font

    col = 4
    for user in users:
        ws2.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        cell = ws2.cell(row=2, column=col, value=user.last_name.upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws2.cell(row=3, column=col, value="ORE").font = subheader_font
        ws2.cell(row=3, column=col).alignment = center
        ws2.cell(row=3, column=col+1, value="GIORNI").font = subheader_font
        ws2.cell(row=3, column=col+1).alignment = center
        col += 3

    for i, project in enumerate(projects):
        data_row = i + 4
        ws2.cell(row=data_row, column=2, value=i+1)
        ws2.cell(row=data_row, column=3, value=project.name).font = Font(size=9)

        col = 4
        for user in users:
            entries = get_entries_for_user(user.id)
            hours = sum(entries.get(project.id, {}).values())
            giorni = math.ceil(hours / 8) if hours > 0 else 0
            if hours > 0:
                ws2.cell(row=data_row, column=col, value=hours).alignment = center
            if giorni > 0:
                ws2.cell(row=data_row, column=col+1, value=giorni).alignment = center
            col += 3

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 5
    ws2.column_dimensions['C'].width = 30
    for i in range(len(users)):
        ws2.column_dimensions[get_column_letter(4 + i*3)].width = 8
        ws2.column_dimensions[get_column_letter(5 + i*3)].width = 8
        ws2.column_dimensions[get_column_letter(6 + i*3)].width = 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"timeflow_export_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )